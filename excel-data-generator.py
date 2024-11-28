import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_formatted_excel(data_df, data_dictionary_df, filename='student_wellbeing_dataset.xlsx'):
    """
    Creates a formatted Excel file with multiple sheets
    """
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write main data
        data_df.to_excel(writer, sheet_name='Data', index=False)
        
        # Write data dictionary
        data_dictionary_df.to_excel(writer, sheet_name='Data Dictionary', index=False)
        
        # Write summary statistics
        summary_stats = pd.DataFrame({
            'Variable': [],
            'Mean': [],
            'Median': [],
            'Std Dev': [],
            'Min': [],
            'Max': []
        })
        
        for col in data_df.select_dtypes(include=[np.number]).columns:
            summary_stats = pd.concat([summary_stats, pd.DataFrame({
                'Variable': [col],
                'Mean': [data_df[col].mean()],
                'Median': [data_df[col].median()],
                'Std Dev': [data_df[col].std()],
                'Min': [data_df[col].min()],
                'Max': [data_df[col].max()]
            })])
        
        summary_stats.to_excel(writer, sheet_name='Summary Statistics', index=False)
        
        # Get the workbook to apply formatting
        workbook = writer.book
        
        # Format Data sheet
        ws = workbook['Data']
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Format Data Dictionary sheet
        ws = workbook['Data Dictionary']
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Format Summary Statistics sheet
        ws = workbook['Summary Statistics']
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            ws.column_dimensions[get_column_letter(col)].width = 15

# Set random seed for reproducibility
np.random.seed(42)

# Number of participants
n_participants = 500

# Generate participant IDs
participant_ids = [f'P{str(i).zfill(3)}' for i in range(1, n_participants + 1)]

# Generate demographic data
age = np.random.normal(20, 2, n_participants).round().astype(int)
age = np.clip(age, 18, 25)

gender = np.random.choice(['Male', 'Female', 'Non-binary'], n_participants, p=[0.45, 0.45, 0.10])

year_of_study = np.random.choice([1, 2, 3, 4], n_participants, p=[0.3, 0.3, 0.25, 0.15])

# Generate physical activity data
weekly_exercise_hours = np.random.gamma(shape=2, scale=2, size=n_participants) * 3
weekly_exercise_hours = weekly_exercise_hours.round(1)

exercise_types = ['Running', 'Gym', 'Team Sports', 'Swimming', 'Cycling', 'Yoga']
primary_exercise = np.random.choice(exercise_types, n_participants)

exercise_intensity = np.random.choice(['Low', 'Moderate', 'High'], n_participants, p=[0.25, 0.5, 0.25])

# Generate mental well-being scores (0-100 scale)
stress_level = np.random.normal(60, 15, n_participants).round(1)
stress_level = np.clip(stress_level, 0, 100)

anxiety_score = np.random.normal(50, 20, n_participants).round(1)
anxiety_score = np.clip(anxiety_score, 0, 100)

life_satisfaction = np.random.normal(70, 15, n_participants).round(1)
life_satisfaction = np.clip(life_satisfaction, 0, 100)

# Generate academic performance data
gpa = np.random.normal(3.2, 0.4, n_participants).round(2)
gpa = np.clip(gpa, 0, 4.0)

study_hours_per_week = np.random.gamma(shape=2, scale=5, size=n_participants).round(1) * 2

# Generate sleep and lifestyle data
average_sleep_hours = np.random.normal(7, 1, n_participants).round(1)
average_sleep_hours = np.clip(average_sleep_hours, 4, 10)

sleep_quality = np.random.choice(['Poor', 'Fair', 'Good', 'Excellent'], n_participants, p=[0.2, 0.3, 0.3, 0.2])

screen_time_hours = np.random.gamma(shape=2, scale=2, size=n_participants).round(1) * 2

# Generate social engagement metrics
social_activities_per_week = np.random.poisson(3, n_participants)
club_membership = np.random.choice(['Yes', 'No'], n_participants, p=[0.4, 0.6])

# Generate intervention group assignment
intervention_group = np.random.choice(['Control', 'Exercise Program', 'Mindfulness Training'], n_participants, p=[0.33, 0.33, 0.34])

# Create time-based measurements (pre and post intervention)
baseline_fitness_score = np.random.normal(70, 15, n_participants).round(1)
baseline_fitness_score = np.clip(baseline_fitness_score, 0, 100)

# Post-intervention scores with effects
effect_size = {'Control': 1.0, 'Exercise Program': 1.2, 'Mindfulness Training': 1.1}
post_fitness_score = baseline_fitness_score * np.array([effect_size[group] for group in intervention_group])
post_fitness_score += np.random.normal(0, 5, n_participants)
post_fitness_score = np.clip(post_fitness_score, 0, 100).round(1)

# Create main DataFrame
data = pd.DataFrame({
    'ParticipantID': participant_ids,
    'Age': age,
    'Gender': gender,
    'YearOfStudy': year_of_study,
    'WeeklyExerciseHours': weekly_exercise_hours,
    'PrimaryExerciseType': primary_exercise,
    'ExerciseIntensity': exercise_intensity,
    'StressLevel': stress_level,
    'AnxietyScore': anxiety_score,
    'LifeSatisfaction': life_satisfaction,
    'GPA': gpa,
    'StudyHoursPerWeek': study_hours_per_week,
    'AverageSleepHours': average_sleep_hours,
    'SleepQuality': sleep_quality,
    'ScreenTimeHours': screen_time_hours,
    'SocialActivitiesPerWeek': social_activities_per_week,
    'ClubMembership': club_membership,
    'InterventionGroup': intervention_group,
    'BaselineFitnessScore': baseline_fitness_score,
    'PostFitnessScore': post_fitness_score
})

# Create data dictionary
data_dictionary = pd.DataFrame({
    'Variable': data.columns,
    'Description': [
        'Unique identifier for each participant',
        'Age in years',
        'Gender identity',
        'Current year of university study',
        'Hours spent exercising per week',
        'Primary form of exercise',
        'Typical exercise intensity level',
        'Perceived stress level (0-100)',
        'Anxiety assessment score (0-100)',
        'Life satisfaction score (0-100)',
        'Grade Point Average (0-4.0)',
        'Hours spent studying per week',
        'Average hours of sleep per night',
        'Subjective sleep quality rating',
        'Daily screen time in hours',
        'Number of social activities per week',
        'Active membership in university clubs',
        'Assigned intervention group',
        'Fitness score at study start (0-100)',
        'Fitness score after intervention (0-100)'
    ],
    'DataType': [str(data[col].dtype) for col in data.columns],
    'PossibleValues': [
        'P001-P500',
        '18-25',
        'Male, Female, Non-binary',
        '1-4',
        '0-30',
        ', '.join(exercise_types),
        'Low, Moderate, High',
        '0-100',
        '0-100',
        '0-100',
        '0-4.0',
        '0-50',
        '4-10',
        'Poor, Fair, Good, Excellent',
        '0-20',
        '0-10',
        'Yes, No',
        'Control, Exercise Program, Mindfulness Training',
        '0-100',
        '0-100'
    ]
})

# Create the formatted Excel file
create_formatted_excel(data, data_dictionary)

print("Excel file 'student_wellbeing_dataset.xlsx' has been created successfully!")